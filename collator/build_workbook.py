"""
build_workbook.py
-----------------
Builds the consolidated Excel workbook from a list of parsed report dicts.

Produces three worksheets:
  1. Executive Summary  — KPI tiles + consolidated one-row-per-system table
  2. Raw Data           — full detail, one row per report run
  3. Pivot Chart        — clustered bar chart by core system

Usage:
    from collator.build_workbook import build_workbook
    build_workbook(all_data, output_path="data/output/report.xlsx")
"""

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from collator.system_map import SYSTEM_MAP, SYSTEM_ORDER

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")
OTHER_FILL  = PatternFill("solid", fgColor="FFEB9C")
ALT_FILL    = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
TITLE_FILL  = PatternFill("solid", fgColor="0D2137")
WHITE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
thin        = Side(border_style="thin", color="BFBFBF")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _consolidated(all_data: list[dict]) -> dict:
    agg = defaultdict(lambda: {"total": 0, "passed": 0, "failed": 0, "other": 0})
    for row in all_data:
        key = SYSTEM_MAP.get(row["system"], row["system"])
        for k in ["total", "passed", "failed", "other"]:
            agg[key][k] += row[k]
    return {s: agg[s] for s in SYSTEM_ORDER if s in agg}


def _grand(all_data: list[dict]) -> dict:
    return {k: sum(r[k] for r in all_data) for k in ["total", "passed", "failed", "other"]}


def _write_system_table(ws, consolidated, grand, start_row=8):
    headers = ["Core System", "Total Tests", "Passed", "Failed", "Other", "Pass Rate %"]
    col_widths = [42, 14, 14, 14, 14, 16]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = WHITE_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[start_row].height = 25

    for i, (system, vals) in enumerate(consolidated.items()):
        r = start_row + 1 + i
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        pct_s = round(vals["passed"] / vals["total"] * 100, 1) if vals["total"] else 0
        for col, val in enumerate([system, vals["total"], vals["passed"], vals["failed"], vals["other"], f"{pct_s}%"], start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = NORMAL_FONT; cell.border = BORDER
            cell.alignment = LEFT if col == 1 else CENTER
            if col == 3:   cell.fill = PASS_FILL
            elif col == 4: cell.fill = FAIL_FILL if vals["failed"] > 0 else PASS_FILL
            elif col == 5: cell.fill = OTHER_FILL if vals["other"] > 0 else WHITE_FILL
            elif col == 6:
                cell.font = Font(name="Calibri", bold=True, size=10)
                cell.fill = PASS_FILL if pct_s >= 90 else (OTHER_FILL if pct_s >= 70 else FAIL_FILL)
            else: cell.fill = fill
        ws.row_dimensions[r].height = 22

    tr = start_row + 1 + len(consolidated)
    for col, val in enumerate(["GRAND TOTAL", grand["total"], grand["passed"], grand["failed"], grand["other"],
                                f"{round(grand['passed']/grand['total']*100,1)}%"], start=1):
        c = ws.cell(row=tr, column=col, value=val)
        c.font = WHITE_FONT; c.fill = HEADER_FILL
        c.alignment = LEFT if col == 1 else CENTER; c.border = BORDER
    ws.row_dimensions[tr].height = 25
    return tr


def build_workbook(all_data: list[dict], output_path: str | None = None) -> str:
    """Build and save the Excel workbook. Returns the output file path."""
    now = datetime.now()
    if not output_path:
        output_path = f"data/output/RDG - Service Assurance - Automation Summary - v1.0 - {now.strftime('%Y%m%d-%H%M')}.xlsx"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    consolidated = _consolidated(all_data)
    grand        = _grand(all_data)
    wb           = Workbook()

    # ── RAW DATA ─────────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Raw Data"
    ws1.merge_cells("A1:I1")
    ws1["A1"].value = "RDG – Service Assurance – Automation Test Execution Summary"
    ws1["A1"].font = TITLE_FONT; ws1["A1"].fill = TITLE_FILL; ws1["A1"].alignment = CENTER
    ws1.row_dimensions[1].height = 35
    ws1.merge_cells("A2:I2")
    ws1["A2"].value = f"Generated: {now.strftime('%d %B %Y %H:%M')}  |  v1.0  |  {len(all_data)} report rows"
    ws1["A2"].font = Font(name="Calibri", color="FFFFFF", italic=True, size=9)
    ws1["A2"].fill = SUBHDR_FILL; ws1["A2"].alignment = CENTER
    ws1.row_dimensions[2].height = 18; ws1.row_dimensions[3].height = 8

    headers   = ["System", "Release / Report Name", "Run Date", "Tool / Framework",
                 "Total # Test Cases", "Passed", "Failed", "Other (Flaky/Skipped)", "Notes"]
    col_widths = [40, 46, 18, 22, 18, 12, 12, 22, 55]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font = WHITE_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[4].height = 30

    prev_system = None; row_num = 5
    for i, row in enumerate(all_data):
        mapped = SYSTEM_MAP.get(row["system"], row["system"])
        if mapped != prev_system:
            ws1.merge_cells(f"A{row_num}:I{row_num}")
            div = ws1.cell(row=row_num, column=1, value=f"── {mapped} ──")
            div.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
            div.fill = SUBHDR_FILL; div.alignment = CENTER
            for col in range(1, 10): ws1.cell(row=row_num, column=col).border = BORDER
            ws1.row_dimensions[row_num].height = 20; row_num += 1
            prev_system = mapped

        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [row["system"], row["release"], row["run_date"], row["tool"],
                  row["total"], row["passed"], row["failed"], row["other"], row["notes"]]
        for col, val in enumerate(values, start=1):
            cell = ws1.cell(row=row_num, column=col, value=val)
            cell.font = NORMAL_FONT; cell.border = BORDER
            cell.alignment = LEFT if col in (1, 2, 4, 9) else CENTER
            if col == 6:   cell.fill = PASS_FILL
            elif col == 7: cell.fill = FAIL_FILL if row["failed"] > 0 else PASS_FILL
            elif col == 8: cell.fill = OTHER_FILL if row["other"] > 0 else WHITE_FILL
            else:          cell.fill = fill
        ws1.row_dimensions[row_num].height = 35; row_num += 1

    ws1.merge_cells(f"A{row_num}:D{row_num}")
    t = ws1.cell(row=row_num, column=1, value="GRAND TOTALS")
    t.font = WHITE_FONT; t.fill = HEADER_FILL; t.alignment = CENTER; t.border = BORDER
    for col, key in [(5, "total"), (6, "passed"), (7, "failed"), (8, "other")]:
        c = ws1.cell(row=row_num, column=col, value=grand[key])
        c.font = WHITE_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
    ws1.cell(row=row_num, column=9).fill = HEADER_FILL
    ws1.cell(row=row_num, column=9).border = BORDER
    ws1.row_dimensions[row_num].height = 25; ws1.freeze_panes = "A5"

    # ── EXECUTIVE SUMMARY ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Executive Summary")
    ws3.merge_cells("A1:F1")
    ws3["A1"].value = "RDG – Service Assurance – Automation Summary"
    ws3["A1"].font = TITLE_FONT; ws3["A1"].fill = TITLE_FILL; ws3["A1"].alignment = CENTER
    ws3.row_dimensions[1].height = 35
    ws3.merge_cells("A2:F2")
    ws3["A2"].value = f"Generated: {now.strftime('%d %B %Y %H:%M')}  |  {len(all_data)} runs  |  {len(consolidated)} core systems"
    ws3["A2"].font = Font(name="Calibri", color="FFFFFF", italic=True, size=9)
    ws3["A2"].fill = SUBHDR_FILL; ws3["A2"].alignment = CENTER
    ws3.row_dimensions[2].height = 18

    for col, (label, val, bg) in enumerate([
        ("TOTAL TESTS", grand["total"],  "1F3864"),
        ("PASSED",      grand["passed"], "375623"),
        ("FAILED",      grand["failed"], "833232"),
        ("OTHER",       grand["other"],  "7F6000"),
    ], start=1):
        ws3.cell(row=4, column=col, value=label).font = WHITE_FONT
        ws3.cell(row=4, column=col).fill = PatternFill("solid", fgColor=bg)
        ws3.cell(row=4, column=col).alignment = CENTER; ws3.cell(row=4, column=col).border = BORDER
        ws3.column_dimensions[get_column_letter(col)].width = 18
        ws3.cell(row=5, column=col, value=val).font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        ws3.cell(row=5, column=col).fill = PatternFill("solid", fgColor=bg)
        ws3.cell(row=5, column=col).alignment = CENTER; ws3.cell(row=5, column=col).border = BORDER
    ws3.row_dimensions[4].height = 22; ws3.row_dimensions[5].height = 35

    pct = round(grand["passed"] / grand["total"] * 100, 1) if grand["total"] else 0
    ws3.merge_cells("E4:F4")
    ws3.cell(row=4, column=5, value="OVERALL PASS RATE").font = WHITE_FONT
    ws3.cell(row=4, column=5).fill = PatternFill("solid", fgColor="1F3864")
    ws3.cell(row=4, column=5).alignment = CENTER; ws3.cell(row=4, column=5).border = BORDER
    ws3.merge_cells("E5:F5")
    ws3.cell(row=5, column=5, value=f"{pct}%").font = Font(name="Calibri", bold=True, size=20, color="FFFFFF")
    ws3.cell(row=5, column=5).fill = PatternFill("solid", fgColor="375623" if pct >= 80 else "833232")
    ws3.cell(row=5, column=5).alignment = CENTER; ws3.cell(row=5, column=5).border = BORDER
    ws3.column_dimensions["E"].width = 18; ws3.column_dimensions["F"].width = 18
    ws3.row_dimensions[7].height = 8

    _write_system_table(ws3, consolidated, grand, start_row=8)
    ws3.freeze_panes = "A9"

    # ── PIVOT CHART ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Pivot Chart")
    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "Test Execution Results by Core System"
    ws2["A1"].font = TITLE_FONT; ws2["A1"].fill = TITLE_FILL; ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells("A2:G2")
    ws2["A2"].value = f"Generated: {now.strftime('%d %B %Y %H:%M')}  |  {len(all_data)} runs  |  {len(consolidated)} systems"
    ws2["A2"].font = Font(name="Calibri", color="FFFFFF", italic=True, size=9)
    ws2["A2"].fill = SUBHDR_FILL; ws2["A2"].alignment = CENTER

    prow = 4
    _write_system_table(ws2, consolidated, grand, start_row=prow)
    data_end = prow + len(consolidated)

    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = "Test Execution Results by Core System"
    chart.y_axis.title = "Test Cases"; chart.style = 10
    chart.width = 40; chart.height = 22
    cats = Reference(ws2, min_col=1, min_row=prow + 1, max_row=data_end)
    for col_idx, color in [(2, "4472C4"), (3, "70AD47"), (4, "FF0000"), (5, "FFC000")]:
        dr = Reference(ws2, min_col=col_idx, min_row=prow, max_row=data_end)
        s = Series(dr, title_from_data=True)
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color
        chart.series.append(s)
    chart.set_categories(cats)
    ws2.add_chart(chart, f"A{data_end + 4}")

    wb._sheets = [ws3, ws1, ws2]
    wb.save(output_path)
    print(f"✅  Workbook saved: {output_path}")
    return output_path
